from flask import *
from flask_cors import CORS
import cv2
import numpy as np
import base64
from model import Model
from database import Database as MyDatabase
import os
import csv
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io

app = Flask(__name__)
CORS(app)
cv2.ocl.setUseOpenCL(False)

# Define Params
detector=cv2.CascadeClassifier('libs/haarcascade_frontalface_default.xml')

emotion_dict = {0: "Angry", 1: "Disgusted", 2: "Fearful", 3: "Happy", 4: "Neutral", 5: "Sad", 6: "Surprised"}

UPLOAD_FOLDER = '/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Create a new CNN Model and Database Model

model = Model()
database = MyDatabase()

# If SQLite database does not exists create one

database.init_database(emotion_dict)

# Routes

@app.route('/questionnaire', methods = ['POST'])
def questionnaire():
    '''
    EXAMPLE JSON
    {
        "history_id" : 1,
        "value" : [4, 5, 7, 8]
    }
    '''
    content = request.json
    database.add_questionnaire(content)

    return jsonify({'status': 'success'})

def decode_image_PIL(image_data):
    # Convert binary image data to bytes
    image_bytes = bytes(image_data)
    
    # Create a PIL (Pillow) image from bytes
    image = PILImage.open(io.BytesIO(image_bytes))
    
    return image

@app.route('/csv', methods = ['GET'])
def excel():
    rows = database.get_csv_data()

    wb = openpyxl.Workbook()
    ws = wb.active

    row_idx = 1
    for row in rows:
        user_id, image_data, emotion_name, probability = row
        ws.cell(row=row_idx, column=1).value = user_id
        image = decode_image_PIL(image_data)

        # Calculate the aspect ratio of the image
        aspect_ratio = image.width / image.height

        # Set the row height based on the aspect ratio and desired width
        desired_width = 100  # Adjust the width as needed
        row_height = int(desired_width / aspect_ratio)
        ws.row_dimensions[row_idx].height = row_height

        image_excel = Image(image)
        ws.add_image(image_excel, 'B{}'.format(row_idx))
        ws.cell(row=row_idx, column=3).value = emotion_name
        ws.cell(row=row_idx, column=4).value = probability
        row_idx = row_idx + 1

    excel_filename = 'output.xlsx'
    wb.save(excel_filename)

    wb.close()

    return send_file('./output.xlsx')



@app.route('/process_image', methods=['POST'])
def process_image():
    if 'image' not in request.files:
        return jsonify({
            'status' : "failed",
            'data' : "No Image Detected !"
        })
    
    global detector
    global model
    file = request.files['image']
    user_id = request.form.get('user_id')

    # Decode image from file
    img = cv2.imdecode(np.frombuffer(file.read(), np.uint8), cv2.IMREAD_COLOR)
    real_img = img.copy()
    # print(img, file=sys.stderr)

    # Face Emotion Recognition testing
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = detector.detectMultiScale(gray,scaleFactor=1.3, minNeighbors=5)

    for (x, y, w, h) in faces:
        cv2.rectangle(img, (x, y-50), (x+w, y+h+10), (255, 0, 0), 2)
        roi_gray = gray[y:y + h, x:x + w]
        cropped_img = np.expand_dims(np.expand_dims(cv2.resize(roi_gray, (48, 48)), -1), 0)
        prediction = model.predict(cropped_img)
        maxindex = int(np.argmax(prediction))
        cv2.putText(img, emotion_dict[maxindex], (x+20, y-60), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2, cv2.LINE_AA)

    if len(faces) == 0:
        return jsonify({
            'status' : "failed",
            'data' : "No Face Detected in the Image!"
        })

    # _, buffer = cv2.imencode('.jpg', img)
    # response = make_response(buffer.tobytes())

    # Check if user_id exists in database
    check_user_id(user_id)

    # Save image to local folder and database
    history_id = database.upload_image(real_img, img, user_id, prediction.flatten())

    # return response
    return get_history_detail_response(history_id)

@app.route('/get_history', methods=['GET'])
def get_history():
    user_id = request.form.get('user_id')
    histories = database.get_history(user_id)
    histories_json = []
    for history in histories:
        history_detail = database.get_highest_prob(history['history_id'])
        histories_json.append({
            'history_id': history['history_id'],
            'emotion_id': history_detail['emotion_id'], 
            'date': history['date']
        })
    
    return jsonify({'status': 'success', 'data': histories_json})


@app.route('/get_history_detail', methods=['GET'])
def get_history_detail():
    history_id = request.form.get('history_id')
    return get_history_detail_response(history_id)
    
def decode_image(image_data):
    nparr = np.frombuffer(image_data, np.uint8)
    image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    return image

def get_history_detail_response(history_id):
    history_details = database.get_history_details(history_id)
    image_row = database.get_image(history_id)

    image = decode_image(image_row[1])

    # encode image as jpeg
    _, encoded_img = cv2.imencode('.jpg', image)
    base64_img = base64.b64encode(encoded_img)

    # decode testing
    # decoded_data = base64.b64decode(base64_img)
    # np_data = np.frombuffer(decoded_data,np.uint8)
    # img = cv2.imdecode(np_data,cv2.IMREAD_UNCHANGED)
    # cv2.imshow('test', img)
    # cv2.waitKey(0)

    history_details_json = []
    for detail in history_details:
        history_details_json.append({
            'emotion_id': detail[1], 
            'probability' : detail[2],
            
        })
    
    return jsonify({'status': 'success', 'data': {
        'base64_img' : str(base64_img),
        'results' : [
            history_details_json
        ]
    }})


def check_user_id(user_id):
    con = database.get_db_connection()
    cur = con.cursor()

    # Check if the user_id exists in the 'users' table
    cur.execute('SELECT * FROM users WHERE user_id = %s', [user_id])
    user = cur.fetchall()

    # If no user_id was found, create a new one
    if len(user) == 0:
        cur.execute('INSERT INTO users (user_id) VALUES (%s)', [user_id])
        con.commit()

    cur.close()
    con.close()
    return

if __name__=='__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port = port)
    # app.run(debug=True,use_reloader=False, port=os.environ.get('FLASK_RUN_PORT'))